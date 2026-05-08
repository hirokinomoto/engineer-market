from pathlib import Path
from datetime import date
import csv
import re

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]

INPUT_FILES = [
    {
        "path": BASE_DIR / "data" / "raw" / "estat" / "job_market_2012_2022.xlsx",
        "source_file": "2012_2022",
        "start_month": date(2015, 1, 1),
        "end_month": date(2022, 12, 1),
    },
    {
        "path": BASE_DIR / "data" / "raw" / "estat" / "job_market_2023_2025.xlsx",
        "source_file": "2023_2025",
        "start_month": date(2023, 1, 1),
        "end_month": date(2025, 12, 1),
    },
]

TARGET_SHEETS = [
    {
        "sheet_name": "第２１表ー１　新規求人（パート含む常用）",
        "metric_name": "新規求人",
        "unit": "人",
    },
    {
        "sheet_name": "第２１表ー２　有効求人（パート含む常用）",
        "metric_name": "有効求人",
        "unit": "人",
    },
    {
        "sheet_name": "第２１表ー７　有効求人倍率（パート含む常用）",
        "metric_name": "有効求人倍率",
        "unit": "倍",
    },
]

OCCUPATION_NAME = "情報処理・通信技術者"
OCCUPATION_ROW = 12

# e-Stat Excelのヘッダー構造
YEAR_ROW = 2
MONTH_ROW = 4

OUTPUT_DIR = BASE_DIR / "data" / "processed"


def normalize_sheet_name(name: str) -> str:
    return name.strip()


def find_sheet_name(workbook, target_sheet_name: str) -> str:
    target_normalized = normalize_sheet_name(target_sheet_name)

    for sheet_name in workbook.sheetnames:
        if normalize_sheet_name(sheet_name) == target_normalized:
            return sheet_name

    raise ValueError(f"対象シートが見つかりません: {target_sheet_name}")


def parse_month_from_headers(year_value, month_value):
    """
    e-Statの月次列から年月を作る。

    実ファイルでは以下の構造：
    - 2行目：西暦年 例）2015年
    - 4行目：月     例）1月
    """
    if year_value is None or month_value is None:
        return None

    year_text = str(year_value)
    month_text = str(month_value)

    year_match = re.search(r"(20\d{2})年", year_text)
    month_match = re.search(r"(\d{1,2})月", month_text)

    if not year_match or not month_match:
        return None

    year = int(year_match.group(1))
    month = int(month_match.group(1))

    if not 1 <= month <= 12:
        return None

    return date(year, month, 1)


def extract_rows_from_sheet(
    ws,
    metric_name: str,
    unit: str,
    source_file: str,
    start_month: date,
    end_month: date,
):
    rows = []

    occupation = ws.cell(row=OCCUPATION_ROW, column=1).value
    if str(occupation).strip() != OCCUPATION_NAME:
        raise ValueError(
            f"想定した職業分類が見つかりません。"
            f" sheet={ws.title}, row={OCCUPATION_ROW}, value={occupation}"
        )

    for col in range(1, ws.max_column + 1):
        year_value = ws.cell(row=YEAR_ROW, column=col).value
        month_value = ws.cell(row=MONTH_ROW, column=col).value

        month = parse_month_from_headers(year_value, month_value)
        if month is None:
            continue

        if not (start_month <= month <= end_month):
            continue

        raw_value = ws.cell(row=OCCUPATION_ROW, column=col).value
        if raw_value is None:
            continue

        rows.append(
            {
                "month": month.isoformat(),
                "occupation_name": OCCUPATION_NAME,
                "metric_name": metric_name,
                "value": raw_value,
                "unit": unit,
                "source_file": source_file,
            }
        )

    return rows


def extract_file(file_config):
    file_path = file_config["path"]

    if not file_path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

    wb = load_workbook(file_path, data_only=True, read_only=True)

    all_rows = []

    for sheet_config in TARGET_SHEETS:
        actual_sheet_name = find_sheet_name(wb, sheet_config["sheet_name"])
        ws = wb[actual_sheet_name]

        rows = extract_rows_from_sheet(
            ws=ws,
            metric_name=sheet_config["metric_name"],
            unit=sheet_config["unit"],
            source_file=file_config["source_file"],
            start_month=file_config["start_month"],
            end_month=file_config["end_month"],
        )

        all_rows.extend(rows)

    return all_rows


def write_csv(rows, output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "month",
        "occupation_name",
        "metric_name",
        "value",
        "unit",
        "source_file",
    ]

    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    for file_config in INPUT_FILES:
        rows = extract_file(file_config)

        output_path = OUTPUT_DIR / f"estat_job_market_{file_config['source_file']}.csv"
        write_csv(rows, output_path)

        print(f"created: {output_path}")
        print(f"rows: {len(rows)}")

        metric_counts = {}
        for row in rows:
            metric_counts[row["metric_name"]] = metric_counts.get(row["metric_name"], 0) + 1

        print("metric counts:")
        for metric_name, count in metric_counts.items():
            print(f"  {metric_name}: {count}")

        print()


if __name__ == "__main__":
    main()