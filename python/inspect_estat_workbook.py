from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]

FILES = [
    BASE_DIR / "data" / "raw" / "estat" / "job_market_2012_2022.xlsx",
    BASE_DIR / "data" / "raw" / "estat" / "job_market_2023_2025.xlsx",
]

TARGET_KEYWORDS = [
    "新規求人",
    "有効求人",
    "有効求人倍率",
    "情報処理・通信技術者",
]


def inspect_workbook(file_path: Path) -> None:
    print("=" * 80)
    print(f"FILE: {file_path.name}")
    print("=" * 80)

    wb = load_workbook(file_path, data_only=True, read_only=True)

    print("\n[Sheet names]")
    for sheet_name in wb.sheetnames:
        print(f"- {sheet_name}")

    print("\n[Keyword search]")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        hits = []
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue

                value_str = str(value)
                for keyword in TARGET_KEYWORDS:
                    if keyword in value_str:
                        hits.append((cell.coordinate, value_str))
                        break

        if hits:
            print(f"\nSheet: {sheet_name}")
            for coordinate, value in hits[:30]:
                print(f"  {coordinate}: {value}")

            if len(hits) > 30:
                print(f"  ... {len(hits) - 30} more hits")


def main() -> None:
    for file_path in FILES:
        if not file_path.exists():
            print(f"File not found: {file_path}")
            continue

        inspect_workbook(file_path)


if __name__ == "__main__":
    main()